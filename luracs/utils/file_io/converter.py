from pathlib import Path
from containers.spectrum_classes import Spectrum, SpectrumData
from containers.roi_classes import ROI, Fit
import openpyxl as pyxl
from .dispatcher import io_dispatcher
from .xml_parser import xml_parser
from .db_parser import db_parser
from itertools import zip_longest
import numpy as np


def spectrum_exporter(parser: xml_parser, file_type: str, new_name: str) -> Spectrum:
    assert file_type in ("csv", "xlsx"), f"{file_type} can not be export target"

    new_spectrum = Spectrum(len(parser.data["foreground"].y_axis), parser.data["name"])

    new_spectrum.set_foreground(parser.data["foreground"])

    bkg = parser.data.get("background")
    if bkg is not None:
        new_spectrum.set_background(bkg)

    peaks = parser.get_rois()
    for roi in peaks:
        new_spectrum.set_roi(roi)


def _dict_to_xlsx(ws, data: dict):
    headers = list(data.keys())

    columns = []
    for v in data.values():
        if isinstance(v, np.ndarray):
            columns.append(v.tolist())
        elif isinstance(v, (list, tuple)):
            columns.append(v)
        else:
            columns.append([v])

    ws.append(headers)

    for row in zip_longest(*columns, fillvalue=""):
        ws.append(row)


def _write_logged_data_to_xlsx(ws, data: dict, include_spectrogram_data: bool):
    headers = []
    columns = []

    for key, values in data.items():
        # ----------------------------
        # SPECTRUM HANDLING
        # ----------------------------
        if key == "spectrum":
            if not include_spectrogram_data:
                continue  # ✅ skip entirely if not requested

            arr = np.asarray(values)

            if arr.ndim != 2:
                raise ValueError(f"spectrum must be 2D, got shape {arr.shape}")

            n_channels = arr.shape[1]

            for ch in range(n_channels):
                headers.append(f"Ch{ch}")
                columns.append(arr[:, ch].tolist())

            continue  # ✅ IMPORTANT: prevent falling into normal handling

        # ----------------------------
        # NORMAL COLUMNS
        # ----------------------------
        headers.append(key)

        # Normalize values safely
        if isinstance(values, np.ndarray):
            values = values.tolist()
        elif isinstance(values, (list, tuple)):
            values = list(values)
        else:
            values = [values]  # ✅ handle scalars safely

        columns.append(values)

    # ----------------------------
    # WRITE HEADER
    # ----------------------------
    ws.append(headers)

    # ----------------------------
    # WRITE ROWS
    # ----------------------------
    for row in zip_longest(*columns, fillvalue=""):
        ws.append(row)


def spectrogram_exporter(
    parser: db_parser,
    file_type: str,
    new_name: str,
    include_spectrogram_data: bool = False,
):
    assert file_type in ("xlsx"), f"{file_type} can not be export target"

    wb = pyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Header"
    _dict_to_xlsx(ws1, parser.get_header())

    ws2 = wb.create_sheet(title="Summary")
    _dict_to_xlsx(ws2, parser.get_summary())

    ws3 = wb.create_sheet(title="Spectrogram")
    _write_logged_data_to_xlsx(
        ws3, parser.get_spectrogram_by_date(), include_spectrogram_data
    )

    wb.save(new_name)
