from __future__ import annotations
import csv
import numpy as np
import openpyxl as pyxl

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from luracs.containers.spectrum_classes import Spectrum
    from pathlib import Path
    from luracs.containers.roi_classes import ROI


class csv_writer:
    @classmethod
    def export_spectrum(cls, spectrum: Spectrum, file_name: Path | str):
        "Export a spectrum to a csv file"
        if isinstance(file_name, str):
            file_name = Path(file_name)
        acc_spectrum = spectrum.get_foreground()
        cps_spectrum = spectrum.get_foreground(cps=True)
        if cps_spectrum is None:
            cps_spectrum = np.zeros_like(acc_spectrum)

        x_axis = spectrum.x_axis

        with open(str(file_name.with_suffix(".csv")), "w", newline="") as csv_file:
            csv_writer = csv.writer(csv_file, dialect=csv.excel)
            csv_writer.writerow(["Energy/Channel [keV]", "Counts", "CPS"])
            csv_writer.writerows(zip(x_axis, acc_spectrum, cps_spectrum))

    @classmethod
    def export_rois(cls, rois: list[ROI], file_name: Path | str):
        "Export the roi data in list of rois to a csv file"
        if isinstance(file_name, str):
            file_name = Path(file_name)

        with open(str(file_name.with_suffix(".csv")), "w", newline="") as csv_file:
            writer = csv.writer(csv_file, dialect=csv.excel)
            writer.writerow(
                [
                    "Alias",
                    "Lower Bound [keV]",
                    "Upper Bound [keV]",
                    "ROI Counts",
                    "Live Time [s]",
                    "Centroid [keV]",
                    "FWHM [keV]",
                    "Amplitude",
                    "Peak Counts",
                    "G",
                    "B",
                    "N",
                    "Nuclide",
                    "Photopeak [keV]",
                    "Intensity [%]",
                ]
            )

            for roi in rois:
                row = [
                    roi.alias,
                    round(roi.roi_bound[0]),
                    round(roi.roi_bound[1]),
                    roi.roi_counts,
                    roi.live_time,
                ]
                if roi.fit is not None:
                    row.extend(
                        [
                            round(roi.fit.mu, 4),
                            round(roi.fit.fwhm, 4),
                            round(roi.fit.A),
                            round(roi.fit.peak_counts),
                            round(roi.fit.G),
                            round(roi.fit.B),
                            round(roi.fit.N),
                        ]
                    )
                else:
                    row.extend([""] * 7)

                if roi.emission is not None:
                    row.extend(
                        [
                            roi.emission.parent_nuclide,
                            roi.emission.energy_keV,
                            roi.emission.intensity_percent,
                        ]
                    )
                else:
                    row.extend([""] * 3)

                writer.writerow(row)


class xlsx_writer:
    @classmethod
    def export_spectrum(
        cls,
        spectrum: Spectrum,
        file_name: str | Path,
        export_spectra: bool = True,
        export_rois: bool = True,
        export_instrument: bool = True,
    ):
        if isinstance(file_name, str):
            file_name = Path(file_name)

        wb = pyxl.Workbook()

        # Remove default sheet once we add our own
        default_sheet = wb.active

        if export_spectra:
            cls._write_spectrum_sheet(wb, spectrum)

        if export_spectra and spectrum.background is not None:
            cls._write_spectrum_background_sheet(wb, spectrum)

        if export_rois and len(spectrum.ROIs) > 0:
            cls._write_roi_sheet(wb, spectrum.ROIs.values())

        if export_instrument and False:
            cls._write_instrument_sheet(wb, spectrum)

        # Remove empty default sheet if other sheets exist
        if len(wb.sheetnames) > 1:
            wb.remove(default_sheet)

        wb.save(str(file_name.with_suffix(".xlsx")))

    @staticmethod
    def _write_spectrum_sheet(wb: pyxl.Workbook, spectrum: Spectrum):
        ws = wb.create_sheet("Spectrum Foreground")

        acc_spectrum = spectrum.get_foreground()
        cps_spectrum = spectrum.get_foreground(cps=True)

        if cps_spectrum is None:
            cps_spectrum = np.zeros_like(acc_spectrum)

        ws.append(["Energy/Channel [keV]", "Counts", "CPS"])

        for row in zip(spectrum.x_axis, acc_spectrum, cps_spectrum):
            ws.append(list(row))

    @staticmethod
    def _write_spectrum_background_sheet(wb: pyxl.Workbook, spectrum: Spectrum):
        ws = wb.create_sheet("Spectrum Background")

        acc_spectrum = spectrum.get_background()
        cps_spectrum = spectrum.get_background(cps=True)

        if cps_spectrum is None:
            cps_spectrum = np.zeros_like(acc_spectrum)

        ws.append(["Energy/Channel [keV]", "Counts", "CPS"])

        for row in zip(spectrum.x_axis, acc_spectrum, cps_spectrum):
            ws.append(list(row))

    @staticmethod
    def _write_roi_sheet(wb: pyxl.Workbook, rois: list[ROI]):
        ws = wb.create_sheet("ROIs")

        ws.append(
            [
                "Alias",
                "Lower Bound [keV]",
                "Upper Bound [keV]",
                "ROI Counts",
                "Live Time [s]",
                "Centroid [keV]",
                "FWHM [keV]",
                "Amplitude",
                "Peak Counts",
                "G",
                "B",
                "N",
                "Nuclide",
                "Photopeak [keV]",
                "Intensity [%]",
            ]
        )

        for roi in rois:
            row = [
                roi.alias,
                round(roi.roi_bound[0]),
                round(roi.roi_bound[1]),
                roi.roi_counts,
                roi.live_time,
            ]

            if roi.fit is not None:
                row.extend(
                    [
                        round(roi.fit.mu, 4),
                        round(roi.fit.fwhm, 4),
                        round(roi.fit.A),
                        round(roi.fit.peak_counts),
                        round(roi.fit.G),
                        round(roi.fit.B),
                        round(roi.fit.N),
                    ]
                )
            else:
                row.extend([None] * 7)

            if roi.emission is not None:
                row.extend(
                    [
                        roi.emission.parent_nuclide,
                        roi.emission.energy_keV,
                        roi.emission.intensity_percent,
                    ]
                )
            else:
                row.extend([None] * 3)
            ws.append(row)

    @staticmethod
    def _write_instrument_sheet(wb: pyxl.Workbook, spectrum: Spectrum):
        ws = wb.create_sheet("Instrument")

        # Adjust these attributes to match your instrument class
        instrument = getattr(spectrum, "instrument", None)
