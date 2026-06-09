import openpyxl as pyxl
from pathlib import Path

from .db_parser import db_parser
from itertools import zip_longest
import numpy as np

from datetime import datetime
from dataclasses import asdict, is_dataclass

from containers.spectrum_classes import Spectrum, SpectrumData


def _dict_to_xlsx(ws, data):
    if is_dataclass(data):
        data = asdict(data)

    headers = list(data.keys())

    columns = []
    for v in data.values():
        if isinstance(v, np.ndarray):
            columns.append(v.tolist())
        elif isinstance(v, (list, tuple)):
            columns.append(v)
        else:
            columns.append([v])

    ws.append(headers)

    for row in zip_longest(*columns, fillvalue=""):
        ws.append(row)


def _write_logged_data_to_xlsx(
    ws,
    data,
    include_spectrogram_data: bool,
):
    if is_dataclass(data):
        data = asdict(data)

    headers = []
    columns = []

    for key, values in data.items():

        # spectra is now the dataclass field name
        if key == "spectra":
            if not include_spectrogram_data:
                continue

            arr = np.asarray(values)

            if arr.ndim != 2:
                raise ValueError(
                    f"spectra must be 2D, got shape {arr.shape}"
                )

            for ch in range(arr.shape[1]):
                headers.append(f"Ch{ch}")
                columns.append(arr[:, ch].tolist())

            continue

        headers.append(key)

        if isinstance(values, np.ndarray):
            values = values.tolist()
        elif isinstance(values, (list, tuple)):
            values = list(values)
        else:
            values = [values]

        columns.append(values)

    ws.append(headers)

    for row in zip_longest(*columns, fillvalue=""):
        ws.append(row)





class db_writer:
    "Provides methods for converting a spectrogram database to other formats."
    @classmethod
    def export_full_xlsx(
        cls,
        parser: db_parser,
        new_name: str,
        include_spectrogram_data: bool = False,
    ):
        wb = pyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Header"
        _dict_to_xlsx(ws1, parser.get_header())

        ws2 = wb.create_sheet(title="Summary")
        _dict_to_xlsx(ws2, parser.get_summary())

        ws3 = wb.create_sheet(title="Spectrogram")
        _write_logged_data_to_xlsx(
            ws3,
            parser.get_spectrogram_by_date(),
            include_spectrogram_data,
        )

        wb.save(new_name)
        
    @classmethod
    def build_spectrum_from_db(
        cls,
        parser: db_parser,
        new_name: str | Path,
        start_time: datetime = None,
        stop_time: datetime = None,
    ) -> Spectrum:

        assert isinstance(start_time, datetime) or start_time is None
        assert isinstance(stop_time, datetime) or stop_time is None

        new_name = Path(new_name)

        header = parser.get_header()
        summary = parser.get_summary()

        new_spectrum = Spectrum(
            header.channels,
            new_name.stem,
            instrument_id=header.device_id,
            device_id=header.device_id,
        )

        if start_time is None and stop_time is None:
            y_axis = summary.total_spectrum
            live_time = summary.total_duration
            start_date = header.created
            end_date = summary.last_update

        else:
            data = parser.get_spectrogram_by_date(
                start_date=start_time,
                end_date=stop_time,
            )

            y_axis = np.sum(data.spectra, axis=0)

            if len(data.timestamps_datetime):
                start_date = data.timestamps_datetime[0]
                end_date = data.timestamps_datetime[-1]
                live_time = (
                    data.timestamps[-1] - data.timestamps[0]
                )
            else:
                start_date = header.created
                end_date = header.created
                live_time = 0

        new_spectrum.set_foreground(
            SpectrumData(
                y_axis=y_axis,
                channels=header.channels,
                total_counts=int(np.sum(y_axis)),
                live_time=live_time,
                real_time=live_time,
                start_date=start_date,
                end_date=end_date,
                spectrum_name=new_name.stem,
                instrument=header.device_id,
            )
        )

        calibration = list(header.calibration)

        if header.concat != 1:
            calibration = [
                c * header.concat**i
                for i, c in enumerate(calibration)
            ]

        new_spectrum.apply_calibration(calibration)

        return new_spectrum